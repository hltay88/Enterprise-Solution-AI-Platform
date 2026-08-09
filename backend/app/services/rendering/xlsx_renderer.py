"""XLSX rendering for BOM deliverables (Sprint 4.4 / ATLAS-050)."""

from __future__ import annotations

from io import BytesIO
from typing import Any


def render_bom_xlsx(
    *,
    title: str,
    status: str,
    sections: list[dict[str, Any]],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    # Cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = title or "BOM"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A2"] = f"Status: {status}"
    cover["A3"] = "DRAFT — NOT APPROVED" if str(status).lower() != "approved" else "APPROVED"
    cover["A5"] = "Pricing is omitted unless present as authoritative approved data (ATLAS-047)."

    by_type = {str(s.get("section_type") or ""): s for s in sections}

    cover_section = by_type.get("cover") or {}
    row = 7
    for item in cover_section.get("content_items") or []:
        cover[f"A{row}"] = str(item.get("text") or "")
        row += 1

    # Line items
    items_sheet = wb.create_sheet("Line Items")
    headers = [
        "Vendor",
        "Model",
        "SKU",
        "Quantity",
        "Unit",
        "Category",
        "Description",
        "Classification",
    ]
    for col, header in enumerate(headers, start=1):
        cell = items_sheet.cell(1, col, header)
        cell.font = Font(bold=True)

    # Merge classifications by model
    class_map: dict[str, str] = {}
    for item in (by_type.get("classification") or {}).get("content_items") or []:
        structured = item.get("structured_data") or {}
        bom_item = structured.get("bom_item") or {}
        key = str(bom_item.get("product_model") or bom_item.get("sku") or "")
        klass = str(structured.get("classification") or "")
        if key:
            class_map[key] = klass

    line_items = (by_type.get("line_items") or {}).get("content_items") or []
    for index, item in enumerate(line_items, start=2):
        structured = item.get("structured_data") or {}
        bom_item = structured.get("bom_item") or {}
        model = str(bom_item.get("product_model") or "")
        items_sheet.cell(index, 1, bom_item.get("vendor") or "")
        items_sheet.cell(index, 2, model)
        items_sheet.cell(index, 3, bom_item.get("sku") or "")
        items_sheet.cell(index, 4, bom_item.get("quantity"))
        items_sheet.cell(index, 5, bom_item.get("unit") or "")
        items_sheet.cell(index, 6, bom_item.get("category") or "")
        items_sheet.cell(index, 7, bom_item.get("description") or "")
        items_sheet.cell(
            index,
            8,
            class_map.get(model) or class_map.get(str(bom_item.get("sku") or "")) or "",
        )

    # Issues
    issues_sheet = wb.create_sheet("Issues")
    issues_sheet["A1"] = "Issue"
    issues_sheet["A1"].font = Font(bold=True)
    for index, item in enumerate(
        (by_type.get("issues") or {}).get("content_items") or [], start=2
    ):
        prefix = "[REVIEW REQUIRED] " if item.get("review_required") else ""
        issues_sheet.cell(index, 1, prefix + str(item.get("text") or ""))

    # Sources
    sources = wb.create_sheet("Sources")
    sources["A1"] = "Source"
    sources["A1"].font = Font(bold=True)
    for index, item in enumerate(
        (by_type.get("sources") or {}).get("content_items") or [], start=2
    ):
        sources.cell(index, 1, str(item.get("text") or ""))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
